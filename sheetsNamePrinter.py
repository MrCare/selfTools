#!python3
'''
Author: Mr.Car
Date: 2024-05-13 11:05:13
使用前需要激活虚拟环境
'''
import fire
from openpyxl import load_workbook

def _format(count, names):
    result = ""
    result += '数量：{} 个'.format(count) + '\n'
    result += '名称：' + '\n'
    for each in names:
        result += each + '\n'
    return result

def process_excel(file_path):
    '''
    计算 excel 中的 sheet 数量并输出各自名称
    '''
    # 读取 Excel 文件
    wb = load_workbook(file_path)

    # 计算数量
    count = len(wb.sheetnames)

    # 列出名称
    # names = df['名称'].tolist()
    names = wb.sheetnames
    print(_format(count, names))
    return "Done!"

if __name__ == '__main__':
    fire.Fire(process_excel)